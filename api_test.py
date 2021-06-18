# -*- coding: utf-8 -*-
# @Author   :   YaMeng
# @File :   api_test.py
# @Software :   PyCharm
# @Time :   2021/6/15 14:03
# @company  :   湖南省零檬信息技术有限公司

# 接口自动化的步骤：
# 1、编写好测试用例，代码自动读取测试用例里的数据  read_data()
# 2、发送接口请求，得到响应结果  -- 实际结果   func()
# 3、执行结果  vs  预期结果
# 4、写入最终的测试结果到测试用例  -- write_data()

import openpyxl
import requests

def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename) # 加载工作簿
    sheet = wb[sheetname] # 找到sheet
    row = sheet.max_row  # 获取最大的行数
    list_1 = []  # 定义一个空列表，用来接收所有的测试用例
    for item in range(2, row+1):  # 取左不取右，左闭右开
        dict_1 = dict(
        id_reg = sheet.cell(row=item, column=1).value, # 取id
        url_reg = sheet.cell(row=item, column=5).value, # 取url
        data_reg = sheet.cell(row=item, column=6).value, # 取data
        expected_reg = sheet.cell(row=item, column=7).value) # 取expected
        list_1.append(dict_1)  # 将一条一条的测试用例都添加到列表里进行保存
    return list_1
def func(url, data , headers={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}):
    res = requests.post(url=url, json=data, headers=headers)
    res_res = res.json()
    return res_res
def write_data(filename, sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)


def execute_function(filename, sheetname):
    res = read_data(filename, sheetname) # 调用读取函数读取注册接口测试用例
    for case in res:
        case_id = case['id_reg']  # 取出用例编号
        case_url = case.get('url_reg')  # 取出接口地址
        case_data = case.get('data_reg')  # 取值请求参数
        case_expect = case['expected_reg']  # 取出预期结果
        case_data = eval(case_data)  # 通过eavl()函数，将取出的字符串格式的data转换为字典格式的data
        case_expect = eval(case_expect) # 转换预期结果
        real_result = func(url=case_url, data=case_data)  # 调用发送函数传入参数
        case_expect_msg = case_expect['msg']  # 预期结果的msg
        real_result_msg = real_result['msg']  # 实际结果的msg
        print('用例编号：{}'.format(case_id))
        print('预期结果为：{}'.format(case_expect_msg))
        print('实际结果为：{}'.format(real_result_msg))
        if case_expect_msg == real_result_msg:
            print('这条用例通过!!')
            final_result = 'pass'  # 设置变量来接收最终的结果传给写入函数
        else:
            print('这条用例不通过!!!')
            final_result = 'false'
        print('*' * 50)
        write_data(filename, sheetname, case_id+1, 8, final_result)

execute_function('test_case_api_no.xlsx','register')
execute_function('test_case_api_no.xlsx','login')

# 写个3个函数和一个判断
# 先读取数据，拿到数据后，执行发送请求，得到预期结果和实际结果
# 做结果判断，然后再把最终的结果写入到测试用例里面




